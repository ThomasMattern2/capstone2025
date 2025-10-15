import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re

def append_data_to_excel(data, filepath):
    """
    Appends parsed MotoCalc data to an Excel file.
    Creates a new workbook if the file does not exist.
    """
    # Create new workbook if file doesn't exist
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "MotoCalc Data"
        # Create headers
        headers = [
            "Motor", "Battery", "Speed Control",
            "AirSpd", "Drag", "Lift"
        ]
        ws.append(headers)
        wb.save(filepath)

    # Load the workbook
    wb = load_workbook(filepath)
    ws = wb.active

    # Append data rows
    for drag_entry, lift_entry in zip(data["Drag Motor"], data["Lift Prop"]):
        ws.append([
            data["Motor"],
            data["Battery"],
            data["Speed Control"],
            drag_entry["AirSpd"],
            drag_entry["Drag"],
            lift_entry["Lift"]
        ])

    wb.save(filepath)
    print(f"✅ Data appended successfully to {filepath}")

def parse_motocalc_file(filepath):
    data = {
        "Motor": "",
        "Battery": "",
        "Speed Control": "",
        "Drag Motor": [],
        "Lift Prop": []
    }

    # Try UTF-8 first, fall back to Windows-1252 if it fails
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f if line.strip()]
    except UnicodeDecodeError:
        with open(filepath, "r", encoding="windows-1252") as f:
            lines = [line.strip() for line in f if line.strip()]

    # --- Extract key-value info (Motor, Battery, etc.) ---
    header_index = None
    for line in lines:
        if line.startswith("Motor:"):
            data["Motor"] = line.split("Motor:")[1].strip()
        elif line.startswith("Drive System:"):
            data["Drive System"] = line.split("Drive System:")[1].strip()
        elif line.startswith("Speed Control:"):
            data["Speed Control"] = line.split("Speed Control:")[1].strip()
        elif re.match(r"^AirSpd", line):
            header_index = lines.index(line)
            break

    # --- Extract table data ---
    if header_index is not None:
        # Skip header + one line of units
        data_lines = lines[header_index + 2:]

        for line in data_lines:
            # Stop before the "Note:" section
            if line.startswith("Note:"):
                break

            # Split numeric values (handles spacing)
            values = re.split(r"\s+", line.strip())

            if len(values) < 3:
                continue

            try:
                airspeed = float(values[0])
                drag = float(values[1])
                lift = float(values[2])
            except ValueError:
                continue

            data["Drag Motor"].append({"AirSpd": airspeed, "Drag": drag})
            data["Lift Prop"].append({"AirSpd": airspeed, "Lift": lift})

    return data



def get_all_txt_files(directory):
    """Return a list of all .txt files in the directory and subdirectories."""
    txt_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.txt'):
                txt_files.append(os.path.join(root, file))
    return txt_files


txt_files = get_all_txt_files(r"C:\Users\jjmik\NonOneDriveDesktop\Y4 S1\Engg 501 Capstone\createdReports")
print(txt_files)

updateExcel = parse_motocalc_file(txt_files[0])
print(parse_motocalc_file(txt_files[0]))

for file in txt_files:
    updateExcel = parse_motocalc_file(file)
    append_data_to_excel(updateExcel, "motocalc_results2.xlsx")